# Copyright (c) 2017-2023 Fumito Hamamura <fumito.ham@gmail.com>

# This library is free software: you can redistribute it and/or
# modify it under the terms of the GNU Lesser General Public
# License as published by the Free Software Foundation version 3.
#
# This library is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the GNU
# Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public
# License along with this library.  If not, see <http://www.gnu.org/licenses/>.

import re
import string
import itertools
import pathlib
import openpyxl as opxl
import openpyxl.cell
from collections.abc import Mapping
from .baseio import BaseIOSpec, BaseSharedIO


def _get_col_index(name):
    """Convert column name to index."""

    index = string.ascii_uppercase.index
    col = 0
    for c in name.upper():
        col = col * 26 + index(c) + 1
    return col


def _is_range_address(range_addr):

    # RANGE_EXPR modified from openpyxl.utils.cells
    # see https://bitbucket.org/openpyxl/openpyxl

    RANGE_EXPR = """
    (?P<cells>
    [$]?(?P<min_col>[A-Za-z]{1,3})?
    [$]?(?P<min_row>\d+)?
    (:[$]?(?P<max_col>[A-Za-z]{1,3})?
    [$]?(?P<max_row>\d+)?)?
    )$
    """
    RANGE_EXPR_RE = re.compile(RANGE_EXPR, re.VERBOSE)

    match = RANGE_EXPR_RE.match(range_addr)

    if not match:
        return False
    else:
        cells = match.group("cells")

    if not cells:
        return False
    else:
        min_col = _get_col_index(match.group("min_col"))
        min_row = int(match.group("min_row"))

        # if range_addr is for a single cell,
        # max_col and max_row are None.
        max_col = match.group("max_col")
        max_col = max_col and _get_col_index(max_col)

        max_row = match.group("max_row")
        max_row = max_row and int(max_row)

        if max_col and max_row:
            return (
                (min_col <= max_col)
                and (min_row <= max_row)
                and (max_col <= 16384)
                and (max_row <= 1048576)
            )
        else:
            return (min_col <= 16384) and (min_row <= 1048576)


def _get_range(book, range_, sheet):
    """Return a range as nested dict of openpyxl cells."""

    filename = None
    if isinstance(book, str):
        filename = book
        book = opxl.load_workbook(book, data_only=True)
    elif isinstance(book, opxl.Workbook):
        pass
    else:
        raise TypeError

    if _is_range_address(range_):
        sheet_names = [name.upper() for name in book.sheetnames]
        index = sheet_names.index(sheet.upper())
        data = book.worksheets[index][range_]
    else:
        data = _get_namedrange(book, range_, sheet)
        if data is None:
            raise ValueError(
                "Named range '%s' not found in %s" % (range_, filename or book)
            )

    return data


def _get_namedrange(book, rangename, sheetname=None):
    """Get range from a workbook.

    A workbook can contain multiple definitions for a single name,
    as a name can be defined for the entire book or for
    a particular sheet.

    If sheet is None, the book-wide def is searched,
    otherwise sheet-local def is looked up.

    Args:
        book: An openpyxl workbook object.
        rangename (str): Range expression, such as "A1", "$G4:$K10",
            named range "NamedRange1".
        sheetname (str, optional): None for book-wide name def,
            sheet name for sheet-local named range.

    Returns:
        Range object specified by the name.

    """
    opxlver = tuple(int(i) for i in opxl.__version__.split('.')[:2])

    if opxlver > (3, 0):

        def cond(namedef):
            if namedef.type.upper() == "RANGE" and namedef.name.upper() == rangename.upper():
                return True
            else:
                return False
    else:
        def cond(namedef):

            if namedef.type.upper() == "RANGE":
                if namedef.name.upper() == rangename.upper():

                    if sheetname is None:
                        if not namedef.localSheetId:
                            return True

                    else:  # sheet local name
                        sheet_id = [sht.upper() for sht in book.sheetnames].index(
                            sheetname.upper()
                        )

                        if namedef.localSheetId == sheet_id:
                            return True

            return False

    def get_destinations(name_def):
        """Workaround for the bug in DefinedName.destinations"""

        from openpyxl.formula import Tokenizer
        from openpyxl.utils.cell import SHEETRANGE_RE

        if name_def.type == "RANGE":
            tok = Tokenizer("=" + name_def.value)
            for part in tok.items:
                if part.subtype == "RANGE":
                    m = SHEETRANGE_RE.match(part.value)
                    if m.group("quoted"):
                        sheet_name = m.group("quoted")
                    else:
                        sheet_name = m.group("notquoted")

                    yield sheet_name, m.group("cells")


    if opxlver > (3, 0):
        # Workbook.defined_names returns DefinedNameDict since openpyxl 3.1.x
        defnamesdict = book[sheetname].defined_names if sheetname else book.defined_names
        namedef = next(
            (item for item in defnamesdict.values() if cond(item)), None
        )
    else:
        # Workbook.defined_names returns DefinedNameList till openpyxl 3.0.x
        namedef = next(
            (item for item in book.defined_names.definedName if cond(item)), None
        )

    if namedef is None:
        return None

    dests = get_destinations(namedef)
    xlranges = []

    sheetnames_upper = [name.upper() for name in book.sheetnames]

    for sht, addr in dests:
        if sheetname:
            sht = sheetname
        index = sheetnames_upper.index(sht.upper())
        xlranges.append(book.worksheets[index][addr])

    if len(xlranges) == 1:
        return xlranges[0]
    else:
        return xlranges


def _redirect_merged(cells):

    if isinstance(cells, openpyxl.cell.Cell):
        return cells
    elif isinstance(cells, openpyxl.cell.MergedCell):
        merged_cells = cells.parent.merged_cells.ranges
        range_ = next((r for r in merged_cells
                       if (r.min_row <= cells.row <= r.max_row)
                       and (r.min_col <= cells.column <= r.max_col)))
        return cells.parent.cell(range_.min_row, range_.min_col)


class ExcelWorkbook(BaseSharedIO):

    def __init__(self, path, manager, load_from):
        from pandas.io.common import IOHandles, get_handle
        # handles = IOHandles(
        #     handle=load_from, compression={"method": None}
        # )
        handles = get_handle(
            load_from, "rb", is_text=False
        )
        super().__init__(path, manager, load_from=load_from)
        self.book = opxl.load_workbook(handles.handle, data_only=True, keep_links=False)

    def _on_write(self, path):
        self.book.save(path)

    def _on_update_value(self, value, kwargs):
        pass

    def get_range(self, range_, sheet):
        return _get_range(self.book, range_, sheet)

    def __getstate__(self):
        state = super().__getstate__()
        state["book"] = self.book
        return state

    def __setstate__(self, state):
        super().__setstate__(state)
        self.book = state["book"]

class _RangeType:

    CELL = 1
    ROW = 2
    COL = COLUMN = 3
    TABLE = 4


class ExcelRange(BaseIOSpec, Mapping):
    """Mapping class for accessing Excel ranges

    An ExcelRange is a dict-like object that
    represents a range in an Excel file.
    The user can read values from the range or write values to it by
    the subscription operator ``[]``.
    ExcelRange is a mapping class, thus it implements all the mapping
    methods and operations.

    ExcelRange objects can only be created by the
    :meth:`Model.new_excel_range<modelx.core.model.Model.new_excel_range>`
    or
    :meth:`UserSpace.new_excel_range<modelx.core.space.UserSpace.new_excel_range>`
    method.

    :class:`ExcelRange` is a subclass of the
    :class:`~modelx.io.baseio.BaseIOSpec` abstract class.
    The :attr:`~modelx.core.model.Model.iospecs` property
    list all the :class:`~modelx.io.baseio.BaseIOSpec` instances
    held in the Model including :class:`ExcelRange` objects.

    See Also:
        * :meth:`UserSpace.new_excel_range<modelx.core.space.UserSpace.new_excel_range>`
        * :meth:`Model.new_excel_range<modelx.core.model.Model.new_excel_range>`
        * :attr:`~modelx.core.model.Model.iospecs`

    .. versionadded:: 0.9.0

    """
    io_class = ExcelWorkbook

    def __init__(self, range_, sheet=None, keyids=None):
        """
        Args:
            path: Path to the Excel file for saving data. If a relative
                path is given, it is relative to the model folder.
            keyids(optional): sequence of strings to specify
                rows and columns to be interpreted as keys.
                E.g. ``["r0", "c1"]`` means keys are pairs of values
                taken from the 1st row and the second column in the ``range_``.
            loadpath(optional): Absolute path to the Excel file to be read in.
                Defaults to ``path``.
        """
        BaseIOSpec.__init__(self)
        self.range = range_
        self.sheet = sheet
        self.keyids = tuple(keyids) if keyids else None

    def _on_load_value(self):
        self._load_cells(self.keyids)

    def _can_update_value(self, value, kwargs):
        return False

    def _on_pickle(self, state):
        state.update({
            "range": self.range,
            "sheet": self.sheet,
            "keyids": self.keyids,
            "_cells": self._cells,
            "_datasize": self._datasize,
            "_key_to_index": self._key_to_index,
            "_keysize": self._keysize
        })
        return state

    def _on_unpickle(self, state):
        self.range = state["range"]
        self.sheet = state["sheet"]
        self.keyids = state["keyids"]
        self._cells = state["_cells"]
        self._datasize = state["_datasize"]
        self._key_to_index = state["_key_to_index"]
        self._keysize = state["_keysize"]

    def _on_serialize(self, state):
        state.update({
            "range": self.range,
            "sheet": self.sheet,
            "keyids": self.keyids
        })
        return state

    def _on_unserialize(self, state):
        self.range = state["range"]
        self.sheet = state["sheet"]
        self.keyids = state["keyids"]
        self._load_cells(self.keyids)

    def _load_cells(self, keys):
        self._cells = self._io.get_range(self.range, self.sheet)
        self._datasize = (len(self._cells), len(self._cells[0]))
        self._key_to_index = self._create_key_to_index(keys)

    def _can_add_other(self, other):
        """Check if self and other have no overlapping cells

            False if:
                self and other on the same sheet
            AND
                overlapping rows
            AND
                overlapping cols
        """
        if self._cells[0][0].parent != other._cells[0][0].parent:
            return True
        elif self._cells[-1][0].row < other._cells[0][0].row:
            return True
        elif other._cells[-1][0].row < self._cells[0][0].row:
            return True
        elif self._cells[0][-1].column < other._cells[0][0].column:
            return True
        elif other._cells[0][-1].column < self._cells[0][0].column:
            return True
        else:
            return False

    def _create_key_to_index(self, keyarg):
        """Initializes self._size, self._keysize and returns a value
        for self._key_to_index"""

        if keyarg is None:
            keyarg = []

        keys = [(p[:1].lower(), int(p[1:])) for p in keyarg]
        key_rows = []
        key_cols = []

        for key in keys:
            if key[0] == "r":
                if key[1] < self._datasize[0]:
                    key_rows.append(key[1])
                else:
                    raise ValueError("invalid row index: %s" % key[1])
            elif key[0] == "c":
                if key[1] < self._datasize[1]:
                    key_cols.append(key[1])
                else:
                    raise ValueError("invalid column index: %s" % key[1])
            else:
                raise ValueError("invalid params: %s" % keyarg)

        for r in key_rows:
            if r < 0 or r >= self._datasize[0]:
                raise ValueError("invalid key row: %s" % r)

        for c in key_cols:
            if c < 0 or c >= self._datasize[0]:
                raise ValueError("invalid key columns: %s" % c)

        self._size = (self._datasize[0] - len(set(key_rows)),
                      self._datasize[1] - len(set(key_cols)))

        rkeys_to_col = {}

        if key_rows:
            for c in range(len(self._cells[0])):
                if c not in key_cols:
                    key = tuple(_redirect_merged(self._cells[r][c]).value
                                for r in key_rows)
                    if key in rkeys_to_col:
                        raise ValueError(
                            "duplicate row key: %s" % repr(key))
                    rkeys_to_col[key] = c

        ckeys_to_row = {}

        if key_cols:
            for r in range(len(self._cells)):
                if r not in key_rows:
                    key = tuple(_redirect_merged(self._cells[r][c]).value
                                for c in key_cols)
                    if key in ckeys_to_row:
                        raise ValueError(
                            "duplicate column key: %s" % repr(key))
                    ckeys_to_row[key] = r

        rkind = [i for i, k in enumerate(keys) if k[0] == "r"]
        ckind = [i for i, k in enumerate(keys) if k[0] == "c"]
        result = {}

        if key_rows and key_cols:

            for rks, cks in itertools.product(rkeys_to_col, ckeys_to_row):
                key = [None] * len(keys)
                for i, k in enumerate(rks):
                    key[rkind[i]] = k
                for i, k in enumerate(cks):
                    key[ckind[i]] = k

                self._keysize = len(key_rows) + len(key_cols)
                result[tuple(key)] = (ckeys_to_row[cks], rkeys_to_col[rks])

        elif key_rows or key_cols:

            if key_rows:
                if self._size[0] > 1:
                    self._keysize = len(key_rows) + 1
                    for rkey, c in rkeys_to_col.items():
                        i = 0
                        for r in range(self._datasize[0]):
                            if r not in key_rows:
                                result[rkey + (i,)] = (r, c)
                                i += 1

                else:
                    self._keysize = len(key_rows)
                    r = next(i for i in range(self._datasize[0])
                             if i not in key_rows)
                    for rkey, c in rkeys_to_col.items():
                        result[rkey] = (r, c)

            else:   # key_cols
                if self._size[1] > 1:
                    self._keysize = len(key_cols) + 1
                    for ckey, r in ckeys_to_row.items():
                        i = 0
                        for c in range(self._datasize[1]):
                            if c not in key_cols:
                                result[ckey + (i,)] = (r, c)
                                i += 1
                else:
                    self._keysize = len(key_cols)
                    c = next(i for i in range(self._datasize[1])
                             if i not in key_cols)
                    for ckey, r in ckeys_to_row.items():
                        result[ckey] = (r, c)

        else:   # No key rows and columns
            if self._datasize[0] > 1 and self._datasize[1] > 1:
                self._keysize = 2

            elif self._datasize[0] > 1 and self._datasize[1] == 1:
                self._keysize = 1

            elif self._datasize[0] == 1 and self._datasize[1] > 1:
                self._keysize = 1

            elif len(self) == 1:
                self._keysize = 0
            else:
                raise RuntimeError("must not happen")

            result = None

        return result

    def _get_index(self, key):

        if self._key_to_index:
            if self._keysize == 1:
                key = (key,)
            return self._key_to_index[key]

        elif self._datasize[0] > 1 and self._datasize[1] > 1:
            return key

        elif self._datasize[0] > 1 and self._datasize[1] == 1:
            return key, 0

        elif self._datasize[0] == 1 and self._datasize[1] > 1:
            return 0, key

        elif len(self) == 1:
            if not bool(key):
                return 0, 0
            else:
                KeyError("invalid key: %s" % repr(key))
        else:
            raise KeyError("invalid key: %s" % repr(key))

    @property
    def value(self):
        return self

    def __getitem__(self, key):
        r, c = self._get_index(key)
        return _redirect_merged(self._cells[r][c]).value

    def __setitem__(self, key, value):
        r, c = self._get_index(key)
        _redirect_merged(self._cells[r][c]).value = value

    def __len__(self):
        return self._size[0] * self._size[1]

    def __iter__(self):
        if self._key_to_index:
            for k in self._key_to_index:
                if self._keysize == 1:
                    yield k[0]
                else:
                    yield k

        elif self._datasize[0] == 1 and self._datasize[1] == 1:
            yield ()
        else:
            if self._datasize[0] > 1 and self._datasize[1] > 1:
                keygen = itertools.product(
                    range(self._datasize[0]), range(self._datasize[1]))
            elif self._datasize[0] > 1 and self._datasize[1] == 1:
                keygen= range(self._datasize[0])
            elif self._datasize[0] == 1 and self._datasize[1] > 1:
                keygen = range(self._datasize[1])
            else:
                raise RuntimeError("must not happen")

            for key in keygen:
                yield key

    def __repr__(self):
        return (
            "<ExcelRange " + "path=%s " + "range=%s " + "sheet=%s>"
        ) % (repr(str(self._io.path.as_posix())),
             repr(self.range), repr(self.sheet))

    def _get_attrdict(self, extattrs=None, recursive=True):
        result = super()._get_attrdict(extattrs=extattrs, recursive=recursive)
        result.update({
            "range": self.range,
            "sheet": self.sheet,
            "keyids": self.keyids})

        return result