# Copyright (c) 2017-2018 Fumito Hamamura <fumito.ham@gmail.com>

# This library is free software: you can redistribute it and/or
# modify it under the terms of the GNU Lesser General Public
# License as published by the Free Software Foundation version 3.
#
# This library is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the GNU
# Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public
# License along with this library.  If not, see <http://www.gnu.org/licenses/>.

import re
import string
import itertools
from collections import namedtuple

import openpyxl as opxl


def _get_col_index(name):
    """Convert column name to index."""

    index = string.ascii_uppercase.index
    col = 0
    for c in name.upper():
        col = col * 26 + index(c) + 1
    return col


def _is_range_address(range_addr):

    # RANGE_EXPR modified from openpyxl.utils.cells
    # see https://bitbucket.org/openpyxl/openpyxl

    RANGE_EXPR = """
    (?P<cells>
    [$]?(?P<min_col>[A-Za-z]{1,3})?
    [$]?(?P<min_row>\d+)?
    (:[$]?(?P<max_col>[A-Za-z]{1,3})?
    [$]?(?P<max_row>\d+)?)?
    )$
    """
    RANGE_EXPR_RE = re.compile(RANGE_EXPR, re.VERBOSE)

    match = RANGE_EXPR_RE.match(range_addr)

    if not match:
        return False
    else:
        cells = match.group("cells")

    if not cells:
        return False
    else:
        min_col = _get_col_index(match.group("min_col"))
        min_row = int(match.group("min_row"))

        # if range_addr is for a single cell,
        # max_col and max_row are None.
        max_col = match.group("max_col")
        max_col = max_col and _get_col_index(max_col)

        max_row = match.group("max_row")
        max_row = max_row and int(max_row)

        if max_col and max_row:
            return ((min_col <= max_col)
                    and (min_row <= max_row)
                    and (max_col <= 16384)
                    and (max_row <= 1048576))
        else:
            return ((min_col <= 16384)
                    and (min_row <= 1048576))


def _get_range(book, range_, sheet):
    """Return a range as nested dict of openpyxl cells."""

    filename = None
    if isinstance(book, str):
        filename = book
        book = opxl.load_workbook(book, data_only=True)
    elif isinstance(book, opxl.Workbook):
        pass
    else:
        raise TypeError

    if _is_range_address(range_):
        sheet_names = [name.upper() for name in book.sheetnames]
        index = sheet_names.index(sheet.upper())
        data = book.worksheets[index][range_]
    else:
        data = _get_namedrange(book, range_, sheet)
        if data is None:
            raise ValueError("Named range '%s' not found in %s" %
                             (range_, filename or book))

    return data


def read_range(filepath, range_expr, sheet=None, dict_generator=None):
    """Read values from an Excel range into a dictionary.

    `range_expr` ie either a range address string, such as "A1", "$C$3:$E$5",
    or a defined name string for a range, such as "NamedRange1".
    If a range address is provided, `sheet` argument must also be provided.
    If a named range is provided and `sheet` is not, book level defined name
    is searched. If `sheet` is also provided, sheet level defined name for the
    specified `sheet` is searched.
    If range_expr points to a single cell, its value is returned.

    `dictgenerator` is a generator function that yields keys and values of 
    the returned dictionary. the excel range, as a nested tuple of openpyxl's
    Cell objects, is passed to the generator function as its single argument.
    If not specified, default generator is used, which maps tuples of row and
    column indexes, both starting with 0, to their values.
    
    Args:
        filepath (str): Path to an Excel file.
        range_epxr (str): Range expression, such as "A1", "$G4:$K10", 
            or named range "NamedRange1"
        sheet (str): Sheet name (case ignored).
            None if book level defined range name is passed as `range_epxr`.
        dict_generator: A generator function taking a nested tuple of cells 
            as a single parameter.

    Returns:
        Nested list containing range values.
    """

    def default_generator(cells):
        for row_ind, row in enumerate(cells):
            for col_ind, cell in enumerate(row):
                yield (row_ind, col_ind), cell.value

    book = opxl.load_workbook(filepath, data_only=True)

    if _is_range_address(range_expr):
        sheet_names = [name.upper() for name in book.sheetnames]
        index = sheet_names.index(sheet.upper())
        cells = book.worksheets[index][range_expr]
    else:
        cells = _get_namedrange(book, range_expr, sheet)

    # In case of a single cell, return its value.
    if isinstance(cells, opxl.cell.Cell):
        return cells.value

    if dict_generator is None:
        dict_generator = default_generator

    gen = dict_generator(cells)
    return {keyval[0]: keyval[1] for keyval in gen}


def _get_namedrange(book, rangename, sheetname=None):
    """Get range from a workbook.

    A workbook can contain multiple definitions for a single name,
    as a name can be defined for the entire book or for
    a particular sheet.

    If sheet is None, the book-wide def is searched,
    otherwise sheet-local def is looked up.

    Args:
        book: An openpyxl workbook object.
        rangename (str): Range expression, such as "A1", "$G4:$K10",
            named range "NamedRange1".
        sheetname (str, optional): None for book-wide name def,
            sheet name for sheet-local named range.

    Returns:
        Range object specified by the name.

    """

    def cond(namedef):

        if namedef.type.upper() == 'RANGE':
            if namedef.name.upper() == rangename.upper():

                if sheetname is None:
                    if not namedef.localSheetId:
                        return True

                else:   # sheet local name
                    sheet_id = [sht.upper() for sht
                                in book.sheetnames].index(sheetname.upper())

                    if namedef.localSheetId == sheet_id:
                        return True

        return False

    def get_destinations(name_def):
        """Workaround for the bug in DefinedName.destinations"""

        from openpyxl.formula import Tokenizer
        from openpyxl.utils.cell import SHEETRANGE_RE

        if name_def.type == "RANGE":
            tok = Tokenizer("=" + name_def.value)
            for part in tok.items:
                if part.subtype == "RANGE":
                    m = SHEETRANGE_RE.match(part.value)
                    if m.group('quoted'):
                        sheet_name = m.group('quoted')
                    else:
                        sheet_name = m.group('notquoted')

                    yield sheet_name, m.group('cells')

    namedef = next((item for item in book.defined_names.definedName
                    if cond(item)), None)

    if namedef is None:
        return None

    dests = get_destinations(namedef)
    xlranges = []

    sheetnames_upper = [name.upper() for name in book.sheetnames]

    for sht, addr in dests:
        if sheetname:
            sht = sheetname
        index = sheetnames_upper.index(sht.upper())
        xlranges.append(book.worksheets[index][addr])

    if len(xlranges) == 1:
        return xlranges[0]
    else:
        return xlranges

_IndexRange = namedtuple('_Index', ['begin', 'len', 'skip'])


class _CellsOrientation:
    ROW = 1
    COL = COLUMN = 2

_ROW = _CellsOrientation.ROW
_COL = _CellsOrientation.COL


class CellsTable:

    def __init__(self, book, range_, sheet,
                 names, params,
                 param_order, transpose,
                 names_ext, params_ext):

        self.data = _get_range(book, range_, sheet)
        self.names_idx = names
        self.param_order = param_order
        self.orientation = _ROW if transpose else _COL

        if self.orientation == _COL:

            self.row_param_cols = params

            if names_ext is not None and params_ext is not None:
                self.col_param_rows = params_ext
            elif names_ext is None and params_ext is None:
                self.col_param_rows = []
            else:
                raise ValueError("invalid pair of name_ext and param_exit")

            self.rows = _IndexRange(0, len(self.data),
                                    [self.names_idx] + self.col_param_rows)
            param_names = []
            for col in self.row_param_cols:
                param_names.append(self.data[self.names_idx][col].value)
            for row in self.col_param_rows:
                param_names.append(self.data[row][names_ext].value)

        elif self.orientation == _ROW:

            self.col_param_rows = params

            if names_ext is not None and params_ext is not None:
                self.row_param_cols = params_ext
            elif names_ext is None and params_ext is None:
                self.row_param_cols = []
            else:
                raise ValueError("invalid pair of name_ext and param_exit")

            self.cols = _IndexRange(0, len(self.data[0]),
                                    [self.names_idx] + self.row_param_cols)
            param_names = []
            for row in self.col_param_rows:
                param_names.append(self.data[row][self.names_idx].value)
            for col in self.row_param_cols:
                param_names.append(self.data[names_ext][col].value)

        else:
            raise ValueError("invalid orientation")

        if self.param_order:
            self.param_names = [param_names[idx] for idx in self.param_order]
        else:
            self.param_order = list(range(len(param_names)))
            self.param_names = param_names

    def items(self):

        if self.orientation == _COL:

            # for each name, yield starting col and length
            name, col_len = None, 0
            for col in range(len(self.data[0])):

                if col in self.row_param_cols:
                    continue

                next_name = self.data[self.names_idx][col].value

                if not next_name:
                    if name:
                        col_len += 1
                    else:
                        raise ValueError("invalid name")
                else:
                    if name is None or name == next_name:
                        name = next_name
                        col_len += 1
                    else:
                        cols_range = _IndexRange(col - col_len, col_len, [])

                        yield _CellsData(self.data, name,
                                         self.row_param_cols,
                                         self.col_param_rows,
                                         self.param_order,
                                         self.rows,
                                         cols_range,
                                         self.orientation)
                        name = next_name
                        col_len = 1

            if col_len > 0:
                cols_range = _IndexRange(len(self.data[0]) - col_len,
                                         col_len, [])

                yield _CellsData(self.data, name,
                                 self.row_param_cols,
                                 self.col_param_rows,
                                 self.param_order,
                                 self.rows,
                                 cols_range,
                                 self.orientation)

        elif self.orientation == _ROW:

            # for each name, yield starting row and length
            name, row_len = None, 0
            for row in range(len(self.data)):

                if row in self.col_param_rows:
                    continue    # skip param row

                next_name = self.data[row][self.names_idx].value

                if not next_name:
                    if name:
                        row_len += 1
                    else:       # first row is blank
                        raise ValueError("invalid name")
                else:
                    if name is None or name == next_name:
                        name = next_name
                        row_len += 1
                    else:
                        rows_range = _IndexRange(row - row_len, row_len, [])

                        yield _CellsData(self.data, name,
                                         self.row_param_cols,
                                         self.col_param_rows,
                                         self.param_order,
                                         rows_range,
                                         self.cols,
                                         self.orientation)
                        name = next_name
                        row_len = 1

            if row_len > 0: # last cells
                rows_range = _IndexRange(len(self.data) - row_len,
                                         row_len, [])

                yield _CellsData(self.data, name,
                                 self.row_param_cols,
                                 self.col_param_rows,
                                 self.param_order,
                                 rows_range,
                                 self.cols,
                                 self.orientation)

        else:
            raise ValueError("invalid orientation")


class _CellsData:

    def __init__(self, data, name,
                 row_param_cols, col_param_rows,
                 param_order,
                 row_range, col_range,
                 orientation):

        self.data = data
        self.name = name
        self.param_order = param_order
        self.row_param_cols = row_param_cols
        self.col_param_rows = col_param_rows
        self.row_range = row_range
        self.col_range = col_range
        self.orientation = orientation

    def params_row(self):

        for row in range(self.row_range.begin,
                         self.row_range.begin + self.row_range.len):

            if row in self.row_range.skip:
                continue

            yield row

    def params_col(self):

        for col in range(self.col_range.begin,
                         self.col_range.begin + self.col_range.len):

            if col in self.col_range.skip:
                continue

            yield col

    def get_param(self, row, col):

        param = []
        for idx in self.param_order:

            if self.orientation == _COL:

                if idx < len(self.row_param_cols):
                    param.append(
                        self.data[row][self.row_param_cols[idx]].value)
                else:
                    idx -= len(self.row_param_cols)
                    param.append(
                        self.data[self.col_param_rows[idx]][col].value)

            elif self.orientation == _ROW:

                if idx < len(self.col_param_rows):
                    param.append(
                        self.data[self.col_param_rows[idx]][col].value)
                else:
                    idx -= len(self.col_param_rows)
                    param.append(
                        self.data[row][self.row_param_cols[idx]].value)

            else:
                raise ValueError('invalid orientation')

        return param

    def params(self):

        for row, col in itertools.product(self.params_row(),
                                          self.params_col()):
            yield self.get_param(row, col)

    def items(self):

        for row, col in itertools.product(self.params_row(),
                                          self.params_col()):
            params = self.get_param(row, col)
            yield params, self.data[row][col].value
